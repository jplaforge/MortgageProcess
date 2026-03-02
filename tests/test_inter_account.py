"""Tests for inter-account transfer detection and reporting."""

import io
from unittest.mock import AsyncMock, MagicMock, patch

import pytest
from openpyxl import load_workbook

from mortgage_mcp.models.bank_statement import DepositCategory
from mortgage_mcp.services.excel_generator import generate_excel
from mortgage_mcp.tools.analyze_bank_statements import _format_summary, analyze_bank_statements


class TestBusinessIncomeExcludesTransfers:
    """Verify that total_business_income excludes personal_transfer deposits."""

    def test_total_excludes_transfers(self, multi_account_extraction):
        """total_business_income should be 40,000 (not 49,500 which includes transfers)."""
        assert multi_account_extraction.total_business_income == 40_000.00
        # Total deposits includes transfers
        assert multi_account_extraction.total_deposits == 49_500.00
        # Difference should be the sum of personal_transfers
        total_transfers = sum(
            m.personal_transfers for m in multi_account_extraction.monthly_breakdown
        )
        assert total_transfers == 9_500.00

    def test_monthly_business_deposits_correct(self, multi_account_extraction):
        """Each month's business_deposits should exclude transfers."""
        months = {m.month: m for m in multi_account_extraction.monthly_breakdown}
        assert months["2025-01"].business_deposits == 14_000.00
        assert months["2025-01"].personal_transfers == 3_500.00
        assert months["2025-02"].business_deposits == 13_000.00
        assert months["2025-02"].personal_transfers == 3_000.00
        assert months["2025-03"].business_deposits == 13_000.00
        assert months["2025-03"].personal_transfers == 3_000.00

    def test_annualized_based_on_business_only(self, multi_account_extraction):
        """Annualized income should be based on business income, not total deposits."""
        expected_monthly_avg = 40_000.00 / 3
        assert multi_account_extraction.average_monthly_business_income == pytest.approx(
            expected_monthly_avg, rel=0.01
        )
        assert multi_account_extraction.annualized_business_income == pytest.approx(
            expected_monthly_avg * 12, rel=0.01
        )


class TestExcelInterAccountTransfers:
    """Verify Excel output reflects inter-account transfers."""

    def test_monthly_sheet_has_transfers_column(self, multi_account_extraction):
        """Detail mensuel sheet column D should contain personal_transfers values."""
        data = generate_excel(multi_account_extraction)
        wb = load_workbook(io.BytesIO(data))
        ws = wb["Detail mensuel"]

        # Column D = "Transferts personnels" (index 4)
        assert ws.cell(row=2, column=4).value == 3_500.00  # Jan
        assert ws.cell(row=3, column=4).value == 3_000.00  # Feb
        assert ws.cell(row=4, column=4).value == 3_000.00  # Mar

    def test_deposits_sheet_shows_personal_transfer_category(self, multi_account_extraction):
        """Depots sheet should have deposits with personal_transfer category."""
        data = generate_excel(multi_account_extraction)
        wb = load_workbook(io.BytesIO(data))
        ws = wb["Depots"]

        transfer_rows = []
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=5).value == "personal_transfer":
                transfer_rows.append(row)

        # 5 inter-account transfers across 3 months
        assert len(transfer_rows) == 5

    def test_deposits_sheet_has_account_column(self, multi_account_extraction):
        """Depots sheet column B should show the account identifier."""
        data = generate_excel(multi_account_extraction)
        wb = load_workbook(io.BytesIO(data))
        ws = wb["Depots"]

        accounts_found = set()
        for row in range(2, ws.max_row + 1):
            account = ws.cell(row=row, column=2).value
            if account:
                accounts_found.add(account)

        assert "Boréale XX89" in accounts_found
        assert "Laurentienne XX77" in accounts_found


class TestTextSummaryInterAccountTransfers:
    """Verify text summary includes transfer information."""

    def test_summary_has_transfers_section(self, multi_account_extraction):
        """Summary should include a 'Transferts inter-comptes' section."""
        summary = _format_summary(multi_account_extraction)
        assert "## Transferts inter-comptes" in summary

    def test_summary_shows_transfer_amount(self, multi_account_extraction):
        """Summary should show the total transfer amount excluded."""
        summary = _format_summary(multi_account_extraction)
        assert "9,500.00 $" in summary

    def test_summary_monthly_table_has_transfers_column(self, multi_account_extraction):
        """Monthly breakdown table should have a Transferts column."""
        summary = _format_summary(multi_account_extraction)
        assert "| Mois | Dépôts affaires | Transferts | Retraits | Nb dépôts |" in summary
        # Check a specific month row contains the transfer amount
        assert "3,500.00 $" in summary  # Jan transfers

    def test_summary_has_transfer_confidence_notes(self, multi_account_extraction):
        """Summary should list transfer-related confidence notes."""
        summary = _format_summary(multi_account_extraction)
        assert "Transfert inter-comptes détecté" in summary
        assert "Boréale XX89" in summary
        assert "Laurentienne XX77" in summary


class TestToolOrchestrationWithTransfers:
    """Verify full pipeline with mocked Vertex AI returns transfer info."""

    @pytest.mark.asyncio
    async def test_full_pipeline_with_transfers(self, multi_account_extraction):
        """Full analyze_bank_statements pipeline should include transfer data in output."""
        import base64

        mock_ctx = MagicMock()
        mock_ctx.report_progress = AsyncMock()
        mock_ctx.info = AsyncMock()

        # Minimal valid PDF (base64)
        fake_pdf_b64 = base64.b64encode(b"%PDF-1.4 fake").decode()

        documents = [
            {"data": fake_pdf_b64, "mime_type": "application/pdf"},
        ]

        with patch(
            "mortgage_mcp.tools.analyze_bank_statements.extract_bank_statements",
            return_value=multi_account_extraction,
        ):
            result = await analyze_bank_statements(
                documents=documents,
                ctx=mock_ctx,
                borrower_name="Martin Girard",
            )

        # Should return TextContent + EmbeddedResource
        assert len(result) == 2

        # Text summary should mention transfers
        text_content = result[0].text
        assert "Transferts inter-comptes" in text_content
        assert "9,500.00 $" in text_content
        assert "40,000.00 $" in text_content  # business income, not 49,500
