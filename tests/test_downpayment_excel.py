"""Tests for downpayment Excel report generation (v4 redesign).

Sheet names (new):
  - "Tableau de bord"  — read-only verdict dashboard (formulas from Analyse)
  - "Analyse"          — main broker work area (dropdowns, running total)
  - "Demandes client"  — document-request tracker with status dropdown
"""

import base64
import io

import pytest
from openpyxl import load_workbook

from mortgage_mcp.models.downpayment import (
    DPAuditResult,
    DPSummary,
    FlagSeverity,
    SourceBreakdown,
)
from mortgage_mcp.services.downpayment_excel import (
    FLAG_TYPE_LABELS,
    SEVERITY_LABELS,
    generate_dp_excel,
    generate_dp_excel_base64,
)


# ── Fixtures ──────────────────────────────────────────────────────────────
# (provided by conftest.py — dp_audit_result)


# ── Core generation ───────────────────────────────────────────────────────

class TestExcelGeneration:
    def test_generates_bytes(self, dp_audit_result):
        data = generate_dp_excel(dp_audit_result)
        assert isinstance(data, bytes)
        assert len(data) > 0

    def test_generates_base64(self, dp_audit_result):
        b64 = generate_dp_excel_base64(dp_audit_result)
        raw = base64.b64decode(b64)
        assert len(raw) > 0

    def test_three_sheets_present(self, dp_audit_result):
        data = generate_dp_excel(dp_audit_result)
        wb = load_workbook(io.BytesIO(data))
        expected = {"Tableau de bord", "Analyse", "Demandes client"}
        assert set(wb.sheetnames) == expected


# ── Tableau de bord ───────────────────────────────────────────────────────

class TestDashboardSheet:
    def _ws(self, result):
        return load_workbook(io.BytesIO(generate_dp_excel(result)))["Tableau de bord"]

    def test_title(self, dp_audit_result):
        ws = self._ws(dp_audit_result)
        assert "Audit de la mise de fonds" in (ws.cell(1, 1).value or "")

    def test_verdict_is_formula_referencing_analyse(self, dp_audit_result):
        """Verdict cell should be a live IF-formula pointing to the Analyse sheet."""
        ws = self._ws(dp_audit_result)
        verdict = ws.cell(3, 1).value or ""
        assert "IF" in str(verdict), "Verdict should be an IF formula"
        assert "Analyse" in str(verdict), "Verdict should reference Analyse sheet"

    def test_verdict_conforme_formula_for_clean_result(self):
        result = DPAuditResult(
            summary=DPSummary(
                dp_target=50000, dp_explained_amount=55000,
                source_breakdown=SourceBreakdown(payroll=55000),
            ),
            borrower_name="Test",
        )
        ws = load_workbook(io.BytesIO(generate_dp_excel(result)))["Tableau de bord"]
        verdict = ws.cell(3, 1).value or ""
        # Formula should contain both possible text values
        assert "CONFORME" in str(verdict)
        assert "À VÉRIFIER" in str(verdict)

    def test_progress_is_formula_with_explique(self, dp_audit_result):
        ws = self._ws(dp_audit_result)
        progress = ws.cell(5, 1).value or ""
        assert "expliqué" in str(progress)
        assert "Analyse" in str(progress)

    def test_borrower_name_present(self, dp_audit_result):
        ws = self._ws(dp_audit_result)
        values = [ws.cell(r, 2).value for r in range(1, 30)]
        assert "Jean Tremblay" in values

    def test_no_co_borrower_when_empty(self, dp_audit_result):
        ws = self._ws(dp_audit_result)
        labels = [ws.cell(r, 1).value for r in range(1, 30)]
        assert "Co-emprunteur :" not in labels

    def test_accounts_section_present(self, dp_audit_result):
        ws = self._ws(dp_audit_result)
        values = [ws.cell(r, 1).value for r in range(1, 50)]
        assert "Comptes analysés" in values
        all_vals = []
        for r in range(1, 50):
            for c in range(1, 7):
                v = ws.cell(r, c).value
                if v:
                    all_vals.append(str(v))
        assert any("Desjardins" in v for v in all_vals)

    def test_zero_sources_hidden(self):
        result = DPAuditResult(
            summary=DPSummary(
                dp_target=50000, dp_explained_amount=20000,
                unexplained_amount=30000, needs_review=True,
                source_breakdown=SourceBreakdown(payroll=20000, gift=0, investment_sale=0),
            ),
            borrower_name="Test",
        )
        ws = load_workbook(io.BytesIO(generate_dp_excel(result)))["Tableau de bord"]
        labels = [ws.cell(r, 1).value for r in range(1, 50)]
        assert "Dons (MCP) :" not in labels
        assert "Vente placement (MCP) :" not in labels
        assert "Accumulation salariale (MCP) :" in labels

    def test_transfers_section_present(self, dp_audit_result):
        ws = self._ws(dp_audit_result)
        values = [ws.cell(r, 1).value for r in range(1, 70)]
        assert "Transferts inter-comptes identifiés" in values

    def test_flags_use_french_labels(self, dp_audit_result):
        ws = self._ws(dp_audit_result)
        all_vals = set()
        for r in range(1, 70):
            for c in range(1, 5):
                v = ws.cell(r, c).value
                if v:
                    all_vals.add(str(v))
        assert any("Dépôt important" in v for v in all_vals)
        assert any("Critique" in v for v in all_vals)
        # No raw English enum values
        assert not any(v == "large_deposit" for v in all_vals)
        assert not any(v == "critical" for v in all_vals)

    def test_legend_present(self, dp_audit_result):
        ws = self._ws(dp_audit_result)
        all_vals = [ws.cell(r, 1).value for r in range(1, 80)]
        assert any(v and "Légende" in str(v) for v in all_vals)

    def test_confirmed_courtier_formula(self, dp_audit_result):
        """'Confirmé courtier' row should reference Analyse inclus total via formula."""
        ws = self._ws(dp_audit_result)
        all_vals = []
        for r in range(1, 50):
            v1 = ws.cell(r, 1).value
            v2 = ws.cell(r, 2).value
            if v1 and "Confirmé courtier" in str(v1):
                assert "Analyse" in str(v2), "Should reference Analyse sheet"
                all_vals.append(v2)
        assert all_vals, "Should have a 'Confirmé courtier' row"


# ── Analyse sheet ─────────────────────────────────────────────────────────

class TestAnalyseSheet:
    def _ws(self, result):
        return load_workbook(io.BytesIO(generate_dp_excel(result)))["Analyse"]

    def test_title(self, dp_audit_result):
        ws = self._ws(dp_audit_result)
        assert "Analyse des dépôts" in (ws.cell(1, 1).value or "")

    def test_instruction_row(self, dp_audit_result):
        ws = self._ws(dp_audit_result)
        instr = ws.cell(2, 1).value or ""
        assert "colonnes jaunes" in instr

    def test_headers_in_row3(self, dp_audit_result):
        ws = self._ws(dp_audit_result)
        headers = [ws.cell(3, c).value for c in range(1, 9)]
        assert "Date" in headers
        assert "Description" in headers
        assert "Montant" in headers
        assert any("Flag" in str(h or "") for h in headers)
        # Broker input column headers (contains ▼)
        assert any("▼" in str(h or "") for h in headers)

    def test_payroll_deposits_pre_set_oui(self, dp_audit_result):
        """Unflagged payroll deposits should start with Inclure=Oui."""
        ws = self._ws(dp_audit_result)
        found_oui = False
        for r in range(4, 30):
            desc = str(ws.cell(r, 2).value or "")
            inclure = ws.cell(r, 6).value
            flag = str(ws.cell(r, 4).value or "")
            if flag == "—" and inclure == "Oui":
                found_oui = True
                break
        assert found_oui, "Unflagged deposits should start with Oui"

    def test_warning_flagged_deposits_start_non(self, dp_audit_result):
        """WARNING-flagged deposits should start with Inclure=Non."""
        ws = self._ws(dp_audit_result)
        for r in range(4, 30):
            flag = str(ws.cell(r, 4).value or "")
            inclure = ws.cell(r, 6).value
            if flag and flag != "—" and "Critique" not in flag:
                # This row has a flag — if it's a warning flag row the fill will be YELLOW
                fill_hex = ws.cell(r, 1).fill.fgColor.rgb if ws.cell(r, 1).fill.fgColor else None
                if fill_hex and "FFEB9C" in str(fill_hex):
                    assert inclure == "Non", f"Warning row should start with Non, got {inclure}"
                    break

    def test_broker_input_columns_have_dropdowns(self, dp_audit_result):
        """Columns E–H should have INPUT_FILL and contain initial values."""
        ws = self._ws(dp_audit_result)
        for r in range(4, 20):
            # Check if it's a data row (has a date in col A)
            if ws.cell(r, 1).value is None:
                continue
            cat_val   = ws.cell(r, 5).value  # E: Catégorie
            incl_val  = ws.cell(r, 6).value  # F: Inclure
            preuve_val = ws.cell(r, 7).value  # G: Preuve
            # All three should be non-empty
            assert cat_val   is not None, f"Row {r}: Catégorie should not be empty"
            assert incl_val  in ("Oui", "Non"), f"Row {r}: Inclure should be Oui or Non, got {incl_val}"
            assert preuve_val is not None, f"Row {r}: Preuve should not be empty"
            break  # Check first data row is enough

    def test_summary_rows_have_formulas(self, dp_audit_result):
        """Summary area should contain SUMIF/SUMIFS formulas."""
        ws = self._ws(dp_audit_result)
        formulas_found = []
        for r in range(1, ws.max_row + 1):
            v = str(ws.cell(r, 3).value or "")
            if "SUMIF" in v or "MAX" in v:
                formulas_found.append(v)
        assert len(formulas_found) >= 2, "Should have SUMIF/MAX formulas in summary"

    def test_currency_formatting_on_montant(self, dp_audit_result):
        ws = self._ws(dp_audit_result)
        for r in range(4, 20):
            if ws.cell(r, 3).value is not None:
                assert "$" in (ws.cell(r, 3).number_format or "")
                break

    def test_institution_names_embedded_in_description(self, dp_audit_result):
        """When multiple accounts, account name should appear in description."""
        ws = self._ws(dp_audit_result)
        all_descs = [str(ws.cell(r, 2).value or "") for r in range(4, 30)]
        # At least one description should contain an institution name
        assert any("Desjardins" in d or "BMO" in d or "TD" in d or "[" in d for d in all_descs)

    def test_flag_column_uses_french(self, dp_audit_result):
        ws = self._ws(dp_audit_result)
        for r in range(4, 20):
            v = ws.cell(r, 4).value
            if v and v != "—":
                assert "large_deposit" not in str(v)
                assert "cash_deposit" not in str(v)
                assert any(
                    french in str(v)
                    for french in ["Dépôt", "Espèces", "Récurrent", "Chaîne", "Couverture",
                                   "Montant", "Succession", "Source", "Crypto", "Devise", "Document"]
                )

    def test_info_flagged_deposits_still_appear(self):
        """INFO-flagged deposits (>=$200) should still appear in Analyse sheet."""
        from mortgage_mcp.models.downpayment import DPFlag, DPTransaction, FlagType, TransactionCategory, TransactionType
        result = DPAuditResult(
            transactions=[
                DPTransaction(id="A1-001", date="2025-01-15", description="SMALL RECURRING",
                              amount=500, type=TransactionType.DEPOSIT,
                              category=TransactionCategory.OTHER, account_id="A1"),
            ],
            flags=[
                DPFlag(type=FlagType.NON_PAYROLL_RECURRING, severity=FlagSeverity.INFO,
                       rationale="test", supporting_transaction_ids=["A1-001"]),
            ],
            summary=DPSummary(dp_target=50000, source_breakdown=SourceBreakdown()),
            borrower_name="Test",
        )
        ws = load_workbook(io.BytesIO(generate_dp_excel(result)))["Analyse"]
        # Should appear in data rows (row 4+) — amount 500 >= threshold
        found = any(ws.cell(r, 3).value == 500 for r in range(4, 20))
        assert found, "INFO-flagged deposit should appear in Analyse sheet"

    def test_empty_result_still_generates(self):
        result = DPAuditResult(
            summary=DPSummary(dp_target=50000, source_breakdown=SourceBreakdown()),
            borrower_name="Test",
        )
        wb = load_workbook(io.BytesIO(generate_dp_excel(result)))
        assert len(wb.sheetnames) == 3


# ── Demandes client ────────────────────────────────────────────────────────

class TestDemandesSheet:
    def _ws(self, result):
        return load_workbook(io.BytesIO(generate_dp_excel(result)))["Demandes client"]

    def test_title(self, dp_audit_result):
        ws = self._ws(dp_audit_result)
        assert "Suivi des demandes" in (ws.cell(1, 1).value or "")

    def test_count_formula_in_row2(self, dp_audit_result):
        ws = self._ws(dp_audit_result)
        v = str(ws.cell(2, 1).value or "")
        assert "COUNTIF" in v, "Row 2 should have a COUNTIF formula"
        assert "À envoyer" in v

    def test_headers_in_row4(self, dp_audit_result):
        ws = self._ws(dp_audit_result)
        headers = [ws.cell(4, c).value for c in range(1, 10)]
        assert "#" in headers
        assert "Statut ▼" in headers
        assert any("Document" in str(h or "") for h in headers)

    def test_request_content_present(self, dp_audit_result):
        ws = self._ws(dp_audit_result)
        all_vals = []
        for r in range(1, 20):
            for c in range(1, 6):
                v = ws.cell(r, c).value
                if v:
                    all_vals.append(str(v))
        assert any("Lettre de don" in v for v in all_vals)

    def test_initial_status_a_envoyer(self, dp_audit_result):
        ws = self._ws(dp_audit_result)
        for r in range(5, 15):
            if ws.cell(r, 1).value is not None:
                assert ws.cell(r, 6).value == "À envoyer"
                break

    def test_transaction_refs_human_readable(self, dp_audit_result):
        """Transaction column should show amount + description, not raw IDs."""
        ws = self._ws(dp_audit_result)
        all_vals = []
        for r in range(5, 20):
            v = ws.cell(r, 5).value
            if v:
                all_vals.append(str(v))
        has_readable = any("25,000" in v or "DON PARENTS" in v for v in all_vals)
        assert has_readable

    def test_docs_have_checkbox_prefix(self, dp_audit_result):
        ws = self._ws(dp_audit_result)
        all_vals = []
        for r in range(1, 20):
            v = ws.cell(r, 4).value  # Documents requis column
            if v:
                all_vals.append(str(v))
        assert any("☐" in v for v in all_vals)

    def test_empty_requests(self):
        result = DPAuditResult(
            summary=DPSummary(dp_target=50000, source_breakdown=SourceBreakdown()),
            borrower_name="Test",
        )
        ws = load_workbook(io.BytesIO(generate_dp_excel(result)))["Demandes client"]
        assert "Aucune demande requise" in (ws.cell(5, 1).value or "")


# ── Flag labels & legacy exports ──────────────────────────────────────────

class TestFlagLabels:
    def test_new_flag_types_have_french_labels(self):
        from mortgage_mcp.models.downpayment import FlagType
        assert FlagType.CRYPTO_SOURCE in FLAG_TYPE_LABELS
        assert FLAG_TYPE_LABELS[FlagType.CRYPTO_SOURCE] == "Source crypto-monnaie"
        assert FlagType.FOREIGN_CURRENCY in FLAG_TYPE_LABELS
        assert FLAG_TYPE_LABELS[FlagType.FOREIGN_CURRENCY] == "Devise étrangère"
        assert FlagType.DOCUMENT_INCOMPLETE in FLAG_TYPE_LABELS
        assert FLAG_TYPE_LABELS[FlagType.DOCUMENT_INCOMPLETE] == "Document incomplet"

    def test_severity_labels_exported(self):
        assert SEVERITY_LABELS[FlagSeverity.CRITICAL] == "Critique"
        assert SEVERITY_LABELS[FlagSeverity.WARNING]  == "Avertissement"
        assert SEVERITY_LABELS[FlagSeverity.INFO]     == "Information"

    def test_broker_input_in_analyse_not_dashboard(self, dp_audit_result):
        """The broker work area should be in Analyse, not Tableau de bord."""
        wb = load_workbook(io.BytesIO(generate_dp_excel(dp_audit_result)))
        ws_db = wb["Tableau de bord"]
        # Dashboard should NOT have the 'Sources à identifier' input header
        db_vals = [ws_db.cell(r, 1).value for r in range(1, 80)]
        assert not any(v and "Sources à identifier" in str(v) for v in db_vals)
        # Analyse SHOULD mention "courtier" in its instruction
        ws_an = wb["Analyse"]
        an_val = ws_an.cell(2, 1).value or ""
        assert "colonnes jaunes" in str(an_val).lower() or "courtier" in str(an_val).lower()

    def test_legend_in_dashboard(self, dp_audit_result):
        ws = load_workbook(io.BytesIO(generate_dp_excel(dp_audit_result)))["Tableau de bord"]
        all_vals = [ws.cell(r, 1).value for r in range(1, 80)]
        assert any(v and "Légende" in str(v) for v in all_vals)


# ── Date formatting ───────────────────────────────────────────────────────

class TestDateFormatting:
    def test_french_date_format(self):
        from mortgage_mcp.services.downpayment_excel import _format_date_short
        assert _format_date_short("2025-02-20") == "20 févr. 2025"
        assert _format_date_short("2025-01-01") == "1 janv. 2025"
        assert _format_date_short("2025-12-25") == "25 déc. 2025"

    def test_invalid_date_passthrough(self):
        from mortgage_mcp.services.downpayment_excel import _format_date_short
        assert _format_date_short("invalid") == "invalid"
        assert _format_date_short("") == ""
