"""Tests for Excel report generation."""

import io

import pytest
from openpyxl import load_workbook

from mortgage_mcp.models.bank_statement import (
    NSFEvent,
    RecurringObligation,
)
from mortgage_mcp.services.excel_generator import generate_excel, generate_excel_base64


class TestGenerateExcel:
    def test_returns_bytes(self, sample_extraction):
        result = generate_excel(sample_extraction)
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_valid_xlsx(self, sample_extraction):
        data = generate_excel(sample_extraction)
        wb = load_workbook(io.BytesIO(data))
        assert "Resume" in wb.sheetnames
        assert "Detail mensuel" in wb.sheetnames
        assert "Depots" in wb.sheetnames
        assert "Retraits" in wb.sheetnames

    def test_resume_sheet_content(self, sample_extraction):
        data = generate_excel(sample_extraction)
        wb = load_workbook(io.BytesIO(data))
        ws = wb["Resume"]

        assert ws["A1"].value == "Grille d'analyse — Revenu de travailleur autonome"
        assert ws["B3"].value == "Jean Tremblay"
        assert ws["B4"].value == "Desjardins"
        assert ws["B7"].value == 3
        # B11 is now a formula referencing Detail mensuel's TOTAL row
        assert "Detail mensuel" in str(ws["B11"].value)
        # B14 is now a formula: B13 * 12
        assert "*12" in str(ws["B14"].value)

    def test_monthly_sheet_rows(self, sample_extraction):
        data = generate_excel(sample_extraction)
        wb = load_workbook(io.BytesIO(data))
        ws = wb["Detail mensuel"]

        # Header row
        assert ws.cell(row=1, column=1).value == "Mois"
        # Data rows: 3 months
        assert ws.cell(row=2, column=1).value == "2025-01"
        assert ws.cell(row=3, column=1).value == "2025-02"
        assert ws.cell(row=4, column=1).value == "2025-03"
        # Col C (Dépôts affaires) is now a SUMPRODUCT formula
        assert "SUMPRODUCT" in str(ws.cell(row=2, column=3).value)
        assert "SUMPRODUCT" in str(ws.cell(row=3, column=3).value)

    def test_monthly_sheet_formulas(self, sample_extraction):
        data = generate_excel(sample_extraction)
        wb = load_workbook(io.BytesIO(data))
        ws = wb["Detail mensuel"]

        # TOTAL row should be at row 5 (3 months + 1 header + 1)
        assert ws.cell(row=5, column=1).value == "TOTAL"
        # Should contain SUM formula
        assert "SUM" in str(ws.cell(row=5, column=2).value)
        # MOYENNE row at 6
        assert ws.cell(row=6, column=1).value == "MOYENNE"
        assert "AVERAGE" in str(ws.cell(row=6, column=2).value)

    def test_deposits_sheet_rows(self, sample_extraction):
        data = generate_excel(sample_extraction)
        wb = load_workbook(io.BytesIO(data))
        ws = wb["Depots"]

        assert ws.cell(row=1, column=1).value == "Date"
        # Total deposits across all months: 4 + 3 + 3 = 10
        total_deposits = sum(
            len(m.deposits) for m in sample_extraction.monthly_breakdown
        )
        # Check last deposit row has data
        assert ws.cell(row=total_deposits + 1, column=1).value is not None

    def test_withdrawals_sheet_rows(self, sample_extraction):
        data = generate_excel(sample_extraction)
        wb = load_workbook(io.BytesIO(data))
        ws = wb["Retraits"]

        # Header
        assert ws.cell(row=1, column=1).value == "Date"
        assert ws.cell(row=1, column=3).value == "Description"
        assert ws.cell(row=1, column=4).value == "Montant"

        # Total withdrawals across all months: 1 (Jan) + 2 (Feb) + 2 (Mar) = 5
        total_withdrawals = sum(
            len(m.withdrawals) for m in sample_extraction.monthly_breakdown
        )
        assert total_withdrawals == 5
        # First withdrawal row
        assert ws.cell(row=2, column=1).value == "2025-01-10"
        assert ws.cell(row=2, column=3).value == "LOYER BUREAU"
        assert ws.cell(row=2, column=4).value == 1500.00
        # Last withdrawal row should have data
        assert ws.cell(row=total_withdrawals + 1, column=1).value is not None

    def test_resume_with_nsf(self, sample_extraction):
        """Resume sheet includes risk indicators when NSF events exist."""
        sample_extraction.nsf_events = [
            NSFEvent(date="2025-01-15", description="NSF CHEQUE 1234", amount=45.00),
            NSFEvent(date="2025-02-20", description="FONDS INSUFFISANTS", amount=45.00),
        ]
        sample_extraction.nsf_total_fees = 90.00

        data = generate_excel(sample_extraction)
        wb = load_workbook(io.BytesIO(data))
        ws = wb["Resume"]

        # Find the risk indicators section
        found_risk = False
        for row in ws.iter_rows(min_col=1, max_col=1):
            if row[0].value == "Indicateurs de risque":
                found_risk = True
                risk_row = row[0].row
                # Next row should show count
                assert ws.cell(row=risk_row + 1, column=2).value == 2
                # Row after that should show total fees
                assert ws.cell(row=risk_row + 2, column=2).value == 90.00
                break
        assert found_risk, "Risk indicators section not found"

    def test_resume_with_obligations(self, sample_extraction):
        """Resume sheet includes recurring obligations when they exist."""
        sample_extraction.recurring_obligations = [
            RecurringObligation(payee="Banque Nationale", monthly_amount=1200.00, category="hypotheque"),
            RecurringObligation(payee="Bell Mobilité", monthly_amount=85.00, category="telecom"),
        ]
        sample_extraction.total_monthly_obligations = 1285.00

        data = generate_excel(sample_extraction)
        wb = load_workbook(io.BytesIO(data))
        ws = wb["Resume"]

        # Find the obligations section
        found_obligations = False
        for row in ws.iter_rows(min_col=1, max_col=1):
            if row[0].value == "Obligations récurrentes détectées":
                found_obligations = True
                section_row = row[0].row
                # Header row
                assert ws.cell(row=section_row + 1, column=1).value == "Bénéficiaire"
                # First obligation
                assert ws.cell(row=section_row + 2, column=1).value == "Banque Nationale"
                assert ws.cell(row=section_row + 2, column=2).value == 1200.00
                # Second obligation
                assert ws.cell(row=section_row + 3, column=1).value == "Bell Mobilité"
                # Total row
                assert ws.cell(row=section_row + 4, column=2).value == 1285.00
                break
        assert found_obligations, "Obligations section not found"

    def test_monthly_transfers_column(self, sample_extraction):
        """Detail mensuel column D should reflect personal_transfers values from each month."""
        data = generate_excel(sample_extraction)
        wb = load_workbook(io.BytesIO(data))
        ws = wb["Detail mensuel"]

        # Column D = "Transferts personnels" (static values)
        assert ws.cell(row=1, column=4).value == "Transferts personnels"
        assert ws.cell(row=2, column=4).value == 2000.00  # Jan
        assert ws.cell(row=3, column=4).value == 1000.00  # Feb
        assert ws.cell(row=4, column=4).value == 1500.00  # Mar

    def test_base64_output(self, sample_extraction):
        result = generate_excel_base64(sample_extraction)
        assert isinstance(result, str)
        # Should be valid base64
        import base64
        decoded = base64.b64decode(result)
        assert decoded[:2] == b"PK"  # ZIP/XLSX magic bytes


class TestBrokerFeatures:
    """Tests for broker-facing improvements."""

    def test_deposits_inclure_column_header(self, sample_extraction):
        """Dépôts sheet should have 'Inclure (O/N)' column F."""
        data = generate_excel(sample_extraction)
        wb = load_workbook(io.BytesIO(data))
        ws = wb["Depots"]

        assert ws.cell(row=1, column=6).value == "Inclure (O/N)"

    def test_deposits_inclure_defaults(self, sample_extraction):
        """Business income deposits default to O; others default to N."""
        data = generate_excel(sample_extraction)
        wb = load_workbook(io.BytesIO(data))
        ws = wb["Depots"]

        for row in range(2, ws.max_row + 1):
            cat = ws.cell(row=row, column=5).value
            inclure = ws.cell(row=row, column=6).value
            if cat is None:
                break
            if cat == "business_income":
                assert inclure == "O", f"Row {row}: business_income should be O"
            else:
                assert inclure == "N", f"Row {row}: {cat} should be N"

    def test_deposits_broker_explanation_column(self, sample_extraction):
        """Dépôts sheet should have 'Explication courtier' column G."""
        data = generate_excel(sample_extraction)
        wb = load_workbook(io.BytesIO(data))
        ws = wb["Depots"]

        assert ws.cell(row=1, column=7).value == "Explication courtier"

    def test_deposits_autofilter(self, sample_extraction):
        """Dépôts sheet should have AutoFilter set."""
        data = generate_excel(sample_extraction)
        wb = load_workbook(io.BytesIO(data))
        ws = wb["Depots"]

        assert ws.auto_filter.ref is not None

    def test_deposits_frozen_pane(self, sample_extraction):
        """Dépôts sheet should have frozen pane at A2."""
        data = generate_excel(sample_extraction)
        wb = load_workbook(io.BytesIO(data))
        ws = wb["Depots"]

        assert ws.freeze_panes == "A2"

    def test_withdrawals_commentaire_column(self, sample_extraction):
        """Retraits sheet should have 'Commentaire courtier' column F."""
        data = generate_excel(sample_extraction)
        wb = load_workbook(io.BytesIO(data))
        ws = wb["Retraits"]

        assert ws.cell(row=1, column=6).value == "Commentaire courtier"

    def test_withdrawals_autofilter(self, sample_extraction):
        """Retraits sheet should have AutoFilter set."""
        data = generate_excel(sample_extraction)
        wb = load_workbook(io.BytesIO(data))
        ws = wb["Retraits"]

        assert ws.auto_filter.ref is not None

    def test_monthly_frozen_pane(self, sample_extraction):
        """Detail mensuel should have frozen pane at A2."""
        data = generate_excel(sample_extraction)
        wb = load_workbook(io.BytesIO(data))
        ws = wb["Detail mensuel"]

        assert ws.freeze_panes == "A2"

    def test_monthly_business_deposits_formula(self, sample_extraction):
        """Detail mensuel col C data rows should use SUMPRODUCT referencing Dépôts."""
        data = generate_excel(sample_extraction)
        wb = load_workbook(io.BytesIO(data))
        ws = wb["Detail mensuel"]

        for row in range(2, len(sample_extraction.monthly_breakdown) + 2):
            val = ws.cell(row=row, column=3).value
            assert "SUMPRODUCT" in str(val), f"Row {row} col C should be SUMPRODUCT formula"
            assert "Depots" in str(val), f"Row {row} col C should reference Depots sheet"

    def test_monthly_revenu_net_formula(self, sample_extraction):
        """Detail mensuel col J data rows should use =C-I formula."""
        data = generate_excel(sample_extraction)
        wb = load_workbook(io.BytesIO(data))
        ws = wb["Detail mensuel"]

        for i, row in enumerate(range(2, len(sample_extraction.monthly_breakdown) + 2), start=2):
            val = ws.cell(row=row, column=10).value
            assert f"=C{row}-I{row}" == val, f"Row {row} col J should be =C{row}-I{row}"

    def test_resume_formula_chain(self, sample_extraction):
        """Resume B11 and B14 should contain formula references."""
        data = generate_excel(sample_extraction)
        wb = load_workbook(io.BytesIO(data))
        ws = wb["Resume"]

        # B11 = Revenu d'affaires total → references Detail mensuel TOTAL row
        assert "Detail mensuel" in str(ws["B11"].value)
        # B13 = Revenu mensuel moyen → references Detail mensuel MOYENNE row
        assert "Detail mensuel" in str(ws["B13"].value)
        # B14 = Revenu annualisé → B13 * 12
        assert "*12" in str(ws["B14"].value)

    def test_resume_broker_sections(self, sample_extraction):
        """Resume should contain broker input sections."""
        data = generate_excel(sample_extraction)
        wb = load_workbook(io.BytesIO(data))
        ws = wb["Resume"]

        all_values = [cell.value for row in ws.iter_rows() for cell in row if cell.value]
        assert any("Informations du dossier" in str(v) for v in all_values)
        assert any("Revenu qualifiable" in str(v) for v in all_values)
        assert any("Attestation du courtier" in str(v) for v in all_values)

    def test_resume_frozen_pane(self, sample_extraction):
        """Resume should have frozen pane at A2."""
        data = generate_excel(sample_extraction)
        wb = load_workbook(io.BytesIO(data))
        ws = wb["Resume"]

        assert ws.freeze_panes == "A2"
