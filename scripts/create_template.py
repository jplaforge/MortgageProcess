#!/usr/bin/env python3
"""Generate the Excel template for income analysis reports.

Run once: python scripts/create_template.py
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

TEMPLATE_DIR = Path(__file__).parent.parent / "src" / "mortgage_mcp" / "templates"
OUTPUT = TEMPLATE_DIR / "grille_revenu_autonome.xlsx"

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def apply_header(ws, row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def main() -> None:
    TEMPLATE_DIR.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # --- Sheet 1: Resume ---
    ws_resume = wb.active
    ws_resume.title = "Resume"

    ws_resume["A1"] = "Grille d'analyse — Revenu de travailleur autonome"
    ws_resume["A1"].font = Font(bold=True, size=14)

    labels_col_a = [
        (3, "Titulaire du compte:"),
        (4, "Institution financière:"),
        (5, "Numéro de compte:"),
        (6, "Période analysée:"),
        (7, "Nombre de mois:"),
    ]
    for row, label in labels_col_a:
        ws_resume.cell(row=row, column=1, value=label).font = Font(bold=True, size=11)

    ws_resume["A9"] = "Sommaire financier"
    ws_resume["A9"].font = Font(bold=True, size=12)

    fin_labels = [
        (10, "Dépôts totaux:"),
        (11, "Revenu d'affaires total:"),
        (12, "Retraits totaux:"),
        (13, "Revenu mensuel moyen (affaires):"),
        (14, "Revenu annualisé (affaires):"),
    ]
    for row, label in fin_labels:
        ws_resume.cell(row=row, column=1, value=label).font = Font(bold=True, size=11)

    ws_resume["A16"] = "Notes et observations"
    ws_resume["A16"].font = Font(bold=True, size=12)

    ws_resume.column_dimensions["A"].width = 35
    ws_resume.column_dimensions["B"].width = 40

    # --- Sheet 2: Detail mensuel ---
    ws_monthly = wb.create_sheet("Detail mensuel")

    monthly_headers = [
        "Mois", "Dépôts bruts", "Dépôts affaires", "Transferts personnels",
        "Gouvernement", "Autres", "Retraits", "Revenu net", "Nb dépôts",
    ]
    for col, header in enumerate(monthly_headers, 1):
        ws_monthly.cell(row=1, column=col, value=header)
    apply_header(ws_monthly, 1, len(monthly_headers))

    widths = [12, 16, 16, 20, 16, 14, 16, 16, 12]
    for i, w in enumerate(widths):
        ws_monthly.column_dimensions[chr(65 + i)].width = w

    # --- Sheet 3: Depots ---
    ws_deposits = wb.create_sheet("Depots")

    deposit_headers = ["Date", "Description", "Montant", "Catégorie"]
    for col, header in enumerate(deposit_headers, 1):
        ws_deposits.cell(row=1, column=col, value=header)
    apply_header(ws_deposits, 1, len(deposit_headers))

    ws_deposits.column_dimensions["A"].width = 14
    ws_deposits.column_dimensions["B"].width = 50
    ws_deposits.column_dimensions["C"].width = 16
    ws_deposits.column_dimensions["D"].width = 20

    wb.save(OUTPUT)
    print(f"Template created: {OUTPUT}")


if __name__ == "__main__":
    main()
