"""Generate the Excel downpayment audit report from analysis results.

Three-sheet workbook designed for mortgage brokers:

1. Tableau de bord  — read-only verdict dashboard with live formulas from Analyse
2. Analyse          — main broker work area: one row per deposit, yellow input columns,
                      running total that feeds the dashboard
3. Demandes client  — document-request tracker with status dropdown & conditional formatting
"""

import base64
import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

from mortgage_mcp.models.downpayment import (
    DPAccountInfo,
    DPAuditResult,
    DPTransaction,
    FlagSeverity,
    FlagType,
    TransactionCategory,
    TransactionType,
)

# ── Palette ────────────────────────────────────────────────────────────────

_BLUE      = "1F4E79"
_GREEN_BG  = "C6EFCE"; _GREEN_FG  = "006100"
_YELLOW_BG = "FFEB9C"; _YELLOW_FG = "9C6500"
_RED_BG    = "FFC7CE"; _RED_FG    = "9C0006"
_BLUE_BG   = "D6EAF8"; _BLUE_FG   = "1F4E79"
_SECT_BG   = "D6E4F0"

HEADER_FILL   = PatternFill(start_color=_BLUE,      end_color=_BLUE,      fill_type="solid")
HEADER_FONT   = Font(bold=True, size=11, color="FFFFFF")
SECTION_FILL  = PatternFill(start_color=_SECT_BG,   end_color=_SECT_BG,   fill_type="solid")
SECTION_FONT  = Font(bold=True, size=11, color=_BLUE)
INPUT_FILL    = PatternFill(start_color="FFFDE7",   end_color="FFFDE7",   fill_type="solid")
INPUT_FONT    = Font(italic=True, color="808080")
ALT_FILL      = PatternFill(start_color="F5F5F5",   end_color="F5F5F5",   fill_type="solid")
GREEN_FILL    = PatternFill(start_color=_GREEN_BG,  end_color=_GREEN_BG,  fill_type="solid")
YELLOW_FILL   = PatternFill(start_color=_YELLOW_BG, end_color=_YELLOW_BG, fill_type="solid")
RED_FILL      = PatternFill(start_color=_RED_BG,    end_color=_RED_BG,    fill_type="solid")
BLUE_FILL     = PatternFill(start_color=_BLUE_BG,   end_color=_BLUE_BG,   fill_type="solid")

THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
CURRENCY = "#,##0.00 $"

SEVERITY_FILLS = {
    FlagSeverity.CRITICAL: RED_FILL,
    FlagSeverity.WARNING:  YELLOW_FILL,
    FlagSeverity.INFO:     BLUE_FILL,
}
SEVERITY_FONTS = {
    FlagSeverity.CRITICAL: Font(color=_RED_FG),
    FlagSeverity.WARNING:  Font(color=_YELLOW_FG),
    FlagSeverity.INFO:     Font(color=_BLUE_FG),
}
SEVERITY_LABELS = {
    FlagSeverity.CRITICAL: "Critique",
    FlagSeverity.WARNING:  "Avertissement",
    FlagSeverity.INFO:     "Information",
}
FLAG_LABELS: dict[FlagType, str] = {
    FlagType.LARGE_DEPOSIT:         "Dépôt important",
    FlagType.CASH_DEPOSIT:          "Dépôt espèces",
    FlagType.NON_PAYROLL_RECURRING: "Récurrent non-salarial",
    FlagType.MULTI_HOP_TRANSFER:    "Chaîne de transferts",
    FlagType.PERIOD_GAP:            "Couverture insuffisante",
    FlagType.ROUND_AMOUNT:          "Montant rond",
    FlagType.RAPID_SUCCESSION:      "Succession rapide",
    FlagType.UNEXPLAINED_SOURCE:    "Source inexpliquée",
    FlagType.CRYPTO_SOURCE:         "Source crypto-monnaie",
    FlagType.FOREIGN_CURRENCY:      "Devise étrangère",
    FlagType.DOCUMENT_INCOMPLETE:   "Document incomplet",
}
CATEGORY_LABELS: dict[TransactionCategory, str] = {
    TransactionCategory.PAYROLL:         "Salaire",
    TransactionCategory.BUSINESS_INCOME: "Revenu d'affaires",
    TransactionCategory.TRANSFER:        "Transfert",
    TransactionCategory.CASH:            "Espèces",
    TransactionCategory.GOVERNMENT:      "Gouvernement",
    TransactionCategory.INVESTMENT:      "Placement",
    TransactionCategory.GIFT:            "Don",
    TransactionCategory.LOAN:            "Prêt",
    TransactionCategory.REFUND:          "Remboursement",
    TransactionCategory.BILL_PAYMENT:    "Facture",
    TransactionCategory.PURCHASE:        "Achat",
    TransactionCategory.OTHER:           "Autre",
}

MONTH_FR = {
    1: "janv.", 2: "févr.", 3: "mars",  4: "avr.",
    5: "mai",   6: "juin",  7: "juill.", 8: "août",
    9: "sept.", 10: "oct.", 11: "nov.", 12: "déc.",
}
_SEV_ORDER = {FlagSeverity.CRITICAL: 0, FlagSeverity.WARNING: 1, FlagSeverity.INFO: 2}

# Source-of-funds categories the broker classifies each deposit into
_SOURCE_OPTIONS = [
    "Salaire accumulé",
    "Virement personnel",
    "Don familial",
    "Vente placement",
    "Revenu d'affaires",
    "Remboursement",
    "Non identifié",
]
_SOURCE_DV = '"' + ",".join(_SOURCE_OPTIONS) + '"'

# Best-guess initial source based on MCP category
_CAT_TO_SOURCE: dict[TransactionCategory, str] = {
    TransactionCategory.PAYROLL:         "Salaire accumulé",
    TransactionCategory.GIFT:            "Don familial",
    TransactionCategory.INVESTMENT:      "Vente placement",
    TransactionCategory.BUSINESS_INCOME: "Revenu d'affaires",
    TransactionCategory.REFUND:          "Remboursement",
    TransactionCategory.TRANSFER:        "Virement personnel",
    TransactionCategory.GOVERNMENT:      "Revenu d'affaires",
}

# Deposits with no flags — start as Oui (confirmed); flagged → Non (needs review)
_DEPOSIT_THRESHOLD = 200.0   # $ minimum to show in Analyse sheet


# ── Helpers ────────────────────────────────────────────────────────────────

def _fmt_date(date_str: str) -> str:
    try:
        dt = datetime.strptime(date_str, "%Y-%m-%d")
        return f"{dt.day} {MONTH_FR[dt.month]} {dt.year}"
    except (ValueError, KeyError):
        return date_str


def _build_acct_lookup(accounts: list[DPAccountInfo]) -> dict[str, str]:
    return {
        a.account_id: (
            f"{a.institution} ({a.account_number_last4})"
            if a.account_number_last4
            else a.institution
        )
        for a in accounts
    }


def _all_transfer_ids(result: DPAuditResult) -> set[str]:
    ids: set[str] = set()
    for m in result.transfers:
        ids.add(m.from_transaction_id)
        if m.to_transaction_id:
            ids.add(m.to_transaction_id)
        ids.update(m.to_transaction_ids)
    return ids


def _header_row(ws, row: int, headers: list[str]) -> None:
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def _section(ws, row: int, title: str, ncols: int = 6) -> int:
    """Write a section-header row spanning ncols. Returns next row."""
    for c in range(1, ncols + 1):
        ws.cell(row=row, column=c).fill = SECTION_FILL
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = SECTION_FONT
    return row + 1


def _set_widths(ws, widths: dict[str, int | float]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def _merge_label_value(
    ws, row: int, label: str, value, bold: bool = False,
    red: bool = False, ncols_label: int = 2, val_col: int = 3,
    num_fmt: str | None = CURRENCY,
) -> None:
    """Write a label (merged) + value pair used in summary blocks."""
    ws.merge_cells(
        start_row=row, start_column=1,
        end_row=row, end_column=ncols_label,
    )
    lbl = ws.cell(row=row, column=1, value=label)
    lbl.font = Font(bold=True, size=11)
    lbl.border = THIN
    val = ws.cell(row=row, column=val_col, value=value)
    val.border = THIN
    if num_fmt:
        val.number_format = num_fmt
    if bold:
        val.font = Font(bold=True, size=12, color=_RED_FG if red else "000000")


# ── Sheet 2: Analyse ──────────────────────────────────────────────────────
# Built before the dashboard so we know the exact summary row addresses.


def _fill_analyse(ws, result: DPAuditResult) -> dict[str, int]:
    """
    Fill the Analyse sheet.
    Returns a dict of row numbers used by the Dashboard for live formulas.
    """
    acct_lookup = _build_acct_lookup(result.accounts)
    transfer_ids = _all_transfer_ids(result)

    # Build per-transaction flag map: id -> (combined_label, highest_severity)
    flagged: dict[str, tuple[str, FlagSeverity]] = {}
    for flag in result.flags:
        label = FLAG_LABELS.get(flag.type, flag.type.value)
        for tid in flag.supporting_transaction_ids:
            if tid not in flagged:
                flagged[tid] = (label, flag.severity)
            else:
                prev_label, prev_sev = flagged[tid]
                combined = f"{prev_label}, {label}"
                best_sev = flag.severity if _SEV_ORDER.get(flag.severity, 3) < _SEV_ORDER.get(prev_sev, 3) else prev_sev
                flagged[tid] = (combined, best_sev)

    # ── Title + instruction row ─────────────────────────────────────────
    ws.merge_cells("A1:H1")
    title = ws.cell(row=1, column=1, value="Analyse des dépôts — Zone de travail du courtier")
    title.font = Font(bold=True, size=14)

    ws.merge_cells("A2:H2")
    instr = ws.cell(
        row=2, column=1,
        value="Complétez les colonnes jaunes (E à H) pour chaque dépôt. "
              "Le Tableau de bord se met à jour automatiquement.",
    )
    instr.fill = INPUT_FILL
    instr.font = Font(italic=True, size=10, color="555555")
    instr.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # ── Column headers (row 3) ──────────────────────────────────────────
    HEADER_ROW = 3
    DATA_START = 4
    headers = [
        "Date", "Description", "Montant",
        "Flag MCP",
        "Catégorie ▼",        # E — broker dropdown
        "Inclure MDF? ▼",     # F — Oui / Non
        "Preuve? ▼",          # G — Reçue / Demandée / Non requise
        "Note courtier",      # H — free text
    ]
    _header_row(ws, HEADER_ROW, headers)
    ws.freeze_panes = "A4"
    ws.row_dimensions[2].height = 24
    ws.row_dimensions[HEADER_ROW].height = 28

    # ── Data validations ────────────────────────────────────────────────
    dv_source = DataValidation(
        type="list", formula1=_SOURCE_DV,
        allow_blank=True, showErrorMessage=False,
    )
    dv_inclure = DataValidation(
        type="list", formula1='"Oui,Non"',
        allow_blank=False, showErrorMessage=False,
    )
    dv_preuve = DataValidation(
        type="list", formula1='"Reçue,Demandée,Non requise"',
        allow_blank=True, showErrorMessage=False,
    )
    for dv in [dv_source, dv_inclure, dv_preuve]:
        ws.add_data_validation(dv)

    # ── Collect deposits to display ────────────────────────────────────
    key_txns: list[tuple[DPTransaction, str | None, FlagSeverity | None]] = []
    for t in result.transactions:
        if t.type != TransactionType.DEPOSIT:
            continue
        if t.id in transfer_ids:
            continue
        flag_label, flag_sev = flagged.get(t.id, (None, None))
        if t.amount >= _DEPOSIT_THRESHOLD or flag_label is not None:
            key_txns.append((t, flag_label, flag_sev))
    key_txns.sort(key=lambda x: x[0].date)

    # ── Data rows ───────────────────────────────────────────────────────
    row = DATA_START
    for i, (txn, flag_label, flag_sev) in enumerate(key_txns):
        acct_suffix = ""
        acct_str = acct_lookup.get(txn.account_id, "")
        if acct_str and len(result.accounts) > 1:
            # Only append account name when there are multiple accounts
            acct_suffix = f"  [{acct_str}]"

        desc = txn.description + acct_suffix

        # Pre-fill broker columns based on MCP analysis
        initial_source = _CAT_TO_SOURCE.get(txn.category, "Non identifié")
        if flag_sev in (FlagSeverity.CRITICAL, FlagSeverity.WARNING):
            initial_inclure = "Non"
            initial_preuve  = "Demandée"
        else:
            initial_inclure = "Oui"
            initial_preuve  = "Non requise"

        # Row background: severity-based for flagged, alternating otherwise
        row_fill: PatternFill | None
        if flag_sev == FlagSeverity.CRITICAL:
            row_fill = RED_FILL
        elif flag_sev == FlagSeverity.WARNING:
            row_fill = YELLOW_FILL
        elif flag_sev == FlagSeverity.INFO:
            row_fill = BLUE_FILL
        elif i % 2 == 1:
            row_fill = ALT_FILL
        else:
            row_fill = None

        # A: Date (read-only)
        ca = ws.cell(row=row, column=1, value=_fmt_date(txn.date))
        ca.border = THIN
        if row_fill:
            ca.fill = row_fill

        # B: Description (read-only)
        cb = ws.cell(row=row, column=2, value=desc)
        cb.border = THIN
        if row_fill:
            cb.fill = row_fill

        # C: Montant (read-only, currency)
        cc = ws.cell(row=row, column=3, value=txn.amount)
        cc.number_format = CURRENCY
        cc.border = THIN
        if row_fill:
            cc.fill = row_fill

        # D: Flag MCP (read-only) — colored by severity
        flag_display = flag_label or "—"
        cd = ws.cell(row=row, column=4, value=flag_display)
        cd.border = THIN
        if flag_sev:
            cd.fill = SEVERITY_FILLS[flag_sev]
            cd.font = SEVERITY_FONTS[flag_sev]
        elif row_fill:
            cd.fill = row_fill

        # E: Catégorie courtier (INPUT)
        ce = ws.cell(row=row, column=5, value=initial_source)
        ce.fill = INPUT_FILL
        ce.border = THIN
        dv_source.add(f"E{row}")

        # F: Inclure MDF (INPUT)
        cf = ws.cell(row=row, column=6, value=initial_inclure)
        cf.fill = INPUT_FILL
        cf.border = THIN
        cf.alignment = Alignment(horizontal="center")
        dv_inclure.add(f"F{row}")

        # G: Preuve (INPUT)
        cg = ws.cell(row=row, column=7, value=initial_preuve)
        cg.fill = INPUT_FILL
        cg.border = THIN
        cg.alignment = Alignment(horizontal="center")
        dv_preuve.add(f"G{row}")

        # H: Note courtier (INPUT, free text)
        ch = ws.cell(row=row, column=8)
        ch.fill = INPUT_FILL
        ch.border = THIN

        row += 1

    last_data_row = max(row - 1, DATA_START)

    # ── Summary block — 4 rows below last data row ─────────────────────
    # These row numbers are returned for Dashboard formula references.
    SUMROW = last_data_row + 2

    f_range = f"$F${DATA_START}:$F${last_data_row}"
    c_range = f"$C${DATA_START}:$C${last_data_row}"
    g_range = f"$G${DATA_START}:$G${last_data_row}"

    inclus_row = SUMROW
    _merge_label_value(
        ws, inclus_row,
        "Total inclus dans la mise de fonds (Oui) :",
        f'=SUMIF({f_range},"Oui",{c_range})',
        bold=True,
    )

    cible_row = SUMROW + 1
    _merge_label_value(
        ws, cible_row,
        "Mise de fonds cible :",
        result.summary.dp_target,
    )

    restant_row = SUMROW + 2
    _merge_label_value(
        ws, restant_row,
        "Montant encore à expliquer :",
        f"=MAX(0,C{cible_row}-C{inclus_row})",
        bold=True, red=True,
    )

    preuve_row = SUMROW + 3
    _merge_label_value(
        ws, preuve_row,
        "Inclus avec preuve reçue :",
        f'=SUMIFS({c_range},{f_range},"Oui",{g_range},"Reçue")',
    )

    # AutoFilter on data columns
    ws.auto_filter.ref = f"A{HEADER_ROW}:H{last_data_row}"

    _set_widths(ws, {
        "A": 16, "B": 48, "C": 16, "D": 26,
        "E": 22, "F": 16, "G": 16, "H": 34,
    })

    return {
        "data_start":    DATA_START,
        "last_data_row": last_data_row,
        "inclus_row":    inclus_row,
        "cible_row":     cible_row,
        "restant_row":   restant_row,
        "preuve_row":    preuve_row,
    }


# ── Sheet 1: Tableau de bord ──────────────────────────────────────────────


def _fill_dashboard(ws, result: DPAuditResult, rows: dict[str, int]) -> None:
    """
    Fill the Tableau de bord sheet.
    Key cells reference Analyse sheet via formulas — verdict updates live.
    """
    acct_lookup = _build_acct_lookup(result.accounts)
    summary = result.summary

    # Shorthand references into Analyse
    restant_ref   = f"Analyse!$C${rows['restant_row']}"
    inclus_ref    = f"Analyse!$C${rows['inclus_row']}"
    cible_ref     = f"Analyse!$C${rows['cible_row']}"
    preuve_ref    = f"Analyse!$C${rows['preuve_row']}"

    row = 1

    # ── Title ──────────────────────────────────────────────────────────
    ws.merge_cells("A1:E1")
    t = ws.cell(row=1, column=1, value="Audit de la mise de fonds — Provenance des fonds")
    t.font = Font(bold=True, size=14)
    row = 3

    # ── Verdict badge (rows 3–4, cols A–E) ────────────────────────────
    VERDICT_ROW = row
    ws.merge_cells(start_row=row, start_column=1, end_row=row + 1, end_column=5)
    verdict_cell = ws.cell(
        row=row, column=1,
        value=f'=IF({restant_ref}<=0,"✅   CONFORME","⚠️   À VÉRIFIER")',
    )
    verdict_cell.font = Font(bold=True, size=18)
    verdict_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 32
    ws.row_dimensions[row + 1].height = 32

    # Conditional formatting: green when restant ≤ 0, yellow otherwise
    verdict_range = f"A{row}:E{row + 1}"
    ws.conditional_formatting.add(
        verdict_range,
        FormulaRule(
            formula=[f"{restant_ref}<=0"],
            fill=GREEN_FILL,
            font=Font(bold=True, size=18, color=_GREEN_FG),
        ),
    )
    ws.conditional_formatting.add(
        verdict_range,
        FormulaRule(
            formula=[f"{restant_ref}>0"],
            fill=YELLOW_FILL,
            font=Font(bold=True, size=18, color=_YELLOW_FG),
        ),
    )

    # Initial fill based on MCP verdict (shows before broker opens Analyse)
    initial_fill = GREEN_FILL if not summary.needs_review else YELLOW_FILL
    for r in range(row, row + 2):
        for c in range(1, 6):
            ws.cell(row=r, column=c).fill = initial_fill
            ws.cell(row=r, column=c).border = THIN
    row += 2

    # ── Progress line ──────────────────────────────────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    prog = ws.cell(
        row=row, column=1,
        value=(
            f'=TEXT({inclus_ref}/{cible_ref},"0%")&" expliqué — "'
            f'&TEXT({inclus_ref},"#,##0 $")&" / "&TEXT({cible_ref},"#,##0 $")'
            f'&"  |  Avec preuve: "&TEXT({preuve_ref},"#,##0 $")'
        ),
    )
    prog.font = Font(size=11)
    prog.alignment = Alignment(horizontal="center")
    row += 2

    # ── Informations du dossier ────────────────────────────────────────
    row = _section(ws, row, "Informations du dossier")
    deal_rows: list[tuple[str, object]] = [("Emprunteur :", result.borrower_name)]
    if result.co_borrower_name:
        deal_rows.append(("Co-emprunteur :", result.co_borrower_name))
    deal_rows.append(("Date de clôture :", _fmt_date(result.closing_date)))
    deal_rows.append(("Mise de fonds cible :", summary.dp_target))
    if result.deal_notes:
        deal_rows.append(("Notes :", result.deal_notes))

    for label, value in deal_rows:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        cell_v = ws.cell(row=row, column=2, value=value)
        if isinstance(value, (int, float)):
            cell_v.number_format = CURRENCY
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        row += 1
    row += 1

    # ── Comptes analysés ───────────────────────────────────────────────
    if result.accounts:
        row = _section(ws, row, "Comptes analysés")
        acct_hdrs = ["Institution", "Titulaire", "No compte", "Période", "Solde ouverture", "Solde fermeture"]
        _header_row(ws, row, acct_hdrs)
        row += 1
        for i, acct in enumerate(result.accounts):
            period = ""
            if acct.period_start and acct.period_end:
                period = f"{_fmt_date(acct.period_start)} — {_fmt_date(acct.period_end)}"
            vals = [
                acct.institution,
                acct.holder_name,
                acct.account_number_last4 or "—",
                period,
                acct.opening_balance,
                acct.closing_balance,
            ]
            fill = ALT_FILL if i % 2 == 1 else None
            for col, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=col, value=v)
                cell.border = THIN
                if col in (5, 6):
                    cell.number_format = CURRENCY
                if fill:
                    cell.fill = fill
            row += 1
        row += 1

    # ── Ventilation — live from Analyse sheet ──────────────────────────
    row = _section(ws, row, "Ventilation confirmée par le courtier  (se met à jour)")
    sb = summary.source_breakdown

    # Static MCP-computed breakdown
    mcp_sources = [
        ("Accumulation salariale (MCP) :",  sb.payroll),
        ("Dons (MCP) :",                    sb.gift),
        ("Vente placement (MCP) :",         sb.investment_sale),
        ("Autres expliquées (MCP) :",       sb.other_explained),
    ]
    for label, val in mcp_sources:
        if val > 0:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=val).number_format = CURRENCY
            row += 1

    # Live courtier-confirmed total (formula)
    ws.cell(row=row, column=1, value="Confirmé courtier :").font = Font(bold=True)
    conf_cell = ws.cell(row=row, column=2, value=f"={inclus_ref}")
    conf_cell.number_format = CURRENCY
    conf_cell.font = Font(bold=True, color=_GREEN_FG)
    row += 1

    ws.cell(row=row, column=1, value="Encore à expliquer :").font = Font(bold=True)
    restant_cell = ws.cell(row=row, column=2, value=f"={restant_ref}")
    restant_cell.number_format = CURRENCY
    restant_cell.font = Font(bold=True, color=_RED_FG)
    row += 2

    # ── Drapeaux MCP ──────────────────────────────────────────────────
    if result.flags:
        n_flags = len(result.flags)
        n_warn  = sum(1 for f in result.flags if f.severity == FlagSeverity.WARNING)
        n_crit  = sum(1 for f in result.flags if f.severity == FlagSeverity.CRITICAL)
        n_info  = sum(1 for f in result.flags if f.severity == FlagSeverity.INFO)
        counts  = []
        if n_crit:  counts.append(f"{n_crit} critique(s)")
        if n_warn:  counts.append(f"{n_warn} avertissement(s)")
        if n_info:  counts.append(f"{n_info} information(s)")

        row = _section(ws, row, f"Drapeaux MCP — {' — '.join(counts)}")
        _header_row(ws, row, ["Type", "Sévérité", "Explication"])
        row += 1

        sorted_flags = sorted(result.flags, key=lambda f: _SEV_ORDER.get(f.severity, 3))
        for i, flag in enumerate(sorted_flags):
            fill = ALT_FILL if i % 2 == 1 else None
            type_cell = ws.cell(row=row, column=1, value=FLAG_LABELS.get(flag.type, flag.type.value))
            type_cell.border = THIN
            if fill:
                type_cell.fill = fill

            sev_cell = ws.cell(row=row, column=2, value=SEVERITY_LABELS.get(flag.severity, ""))
            sev_cell.border = THIN
            sev_cell.fill = SEVERITY_FILLS.get(flag.severity, PatternFill())
            sev_cell.font = SEVERITY_FONTS.get(flag.severity, Font())
            sev_cell.alignment = Alignment(horizontal="center")

            rat_cell = ws.cell(row=row, column=3, value=flag.rationale)
            rat_cell.border = THIN
            if fill:
                rat_cell.fill = fill
            row += 1
        row += 1

    # ── Transfers ─────────────────────────────────────────────────────
    if result.transfers:
        row = _section(ws, row, "Transferts inter-comptes identifiés")
        _header_row(ws, row, ["De", "Vers", "Montant", "Détail"])
        row += 1
        for i, tm in enumerate(result.transfers):
            from_label = acct_lookup.get(tm.from_account_id, tm.from_account_id)
            to_label   = acct_lookup.get(tm.to_account_id,   tm.to_account_id)
            detail = (
                f"Fractionné × {len(tm.to_transaction_ids)}" if tm.is_split
                else (f"Délai: {tm.date_delta_days} j" if tm.date_delta_days > 0 else "Même jour")
            )
            fill = ALT_FILL if i % 2 == 1 else None
            for col, val in enumerate([from_label, to_label, tm.amount, detail], 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = THIN
                if col == 3:
                    cell.number_format = CURRENCY
                if fill:
                    cell.fill = fill
            row += 1
        row += 1

    # ── Legend ────────────────────────────────────────────────────────
    row = _section(ws, row, "Légende", ncols=3)
    legend = [
        ("Critique",        RED_FILL,    Font(color=_RED_FG),    "Action immédiate requise"),
        ("Avertissement",   YELLOW_FILL, Font(color=_YELLOW_FG), "Vérification requise"),
        ("Information",     BLUE_FILL,   Font(color=_BLUE_FG),   "Pour information seulement"),
        ("Zone de saisie",  INPUT_FILL,  INPUT_FONT,             "À compléter par le courtier (feuille Analyse)"),
    ]
    for label, fill, font, desc in legend:
        ca = ws.cell(row=row, column=1, value=label)
        ca.fill = fill; ca.font = font; ca.border = THIN
        cb = ws.cell(row=row, column=2, value=desc)
        cb.border = THIN
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        row += 1

    ws.freeze_panes = "A2"
    _set_widths(ws, {"A": 32, "B": 28, "C": 22, "D": 22, "E": 18, "F": 18})


# ── Sheet 3: Demandes client ──────────────────────────────────────────────


def _fill_demandes(ws, result: DPAuditResult) -> None:
    """
    Fill the Demandes client sheet as a proper tracking table.
    Each client request occupies ONE row; broker updates Statut via dropdown.
    Conditional formatting colors the row by status.
    """
    acct_lookup = _build_acct_lookup(result.accounts)
    tx_by_id = {t.id: t for t in result.transactions}

    # ── Title + pending count formula ──────────────────────────────────
    ws.merge_cells("A1:I1")
    t = ws.cell(row=1, column=1, value="Suivi des demandes au client")
    t.font = Font(bold=True, size=14)

    ws.merge_cells("A2:I2")
    count_cell = ws.cell(
        row=2, column=1,
        value=(
            '=COUNTIF($F$5:$F$200,"À envoyer")&" à envoyer — "'
            '&COUNTIF($F$5:$F$200,"Envoyée")&" en attente de réponse"'
        ),
    )
    count_cell.font = Font(size=11, bold=True)
    count_cell.alignment = Alignment(horizontal="left", indent=1)

    # ── Column headers (row 4) ─────────────────────────────────────────
    HDR_ROW = 4
    DATA_START = 5
    headers = [
        "#",
        "Document demandé",
        "Raison",
        "Documents requis",
        "Transactions concernées",
        "Statut ▼",
        "Date envoyée",
        "Date reçue",
        "Note courtier",
    ]
    _header_row(ws, HDR_ROW, headers)
    ws.freeze_panes = "A5"
    ws.row_dimensions[HDR_ROW].height = 28

    # Status dropdown
    dv_status = DataValidation(
        type="list",
        formula1='"À envoyer,Envoyée,Reçue,Non requis"',
        allow_blank=False,
        showErrorMessage=False,
    )
    ws.add_data_validation(dv_status)

    if not result.client_requests:
        ws.cell(row=DATA_START, column=1, value="Aucune demande requise.").font = Font(italic=True)
        _set_widths(ws, {"A": 12, "B": 40, "C": 40})
        return

    # ── One row per client request ─────────────────────────────────────
    row = DATA_START
    for i, req in enumerate(result.client_requests, 1):
        # Build bullet list of docs required (newline-separated in one cell)
        docs_text = "\n".join(f"☐ {d}" for d in req.required_docs)

        # Build transaction references
        tx_refs = []
        for tid in (req.supporting_transaction_ids or []):
            t = tx_by_id.get(tid)
            if t:
                acct = acct_lookup.get(t.account_id, "")
                tx_refs.append(f"{t.amount:,.0f} $ — {t.description} ({_fmt_date(t.date)})")
        txns_text = "\n".join(tx_refs) if tx_refs else "—"

        row_data = [i, req.title, req.reason, docs_text, txns_text]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = THIN
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        # F: Statut (dropdown, INPUT)
        cell_f = ws.cell(row=row, column=6, value="À envoyer")
        cell_f.fill = INPUT_FILL
        cell_f.border = THIN
        cell_f.alignment = Alignment(horizontal="center", vertical="top")
        dv_status.add(f"F{row}")

        # G: Date envoyée (INPUT)
        cg = ws.cell(row=row, column=7)
        cg.fill = INPUT_FILL; cg.border = THIN; cg.number_format = "YYYY-MM-DD"

        # H: Date reçue (INPUT)
        ch = ws.cell(row=row, column=8)
        ch.fill = INPUT_FILL; ch.border = THIN; ch.number_format = "YYYY-MM-DD"

        # I: Note courtier (INPUT)
        ci = ws.cell(row=row, column=9)
        ci.fill = INPUT_FILL; ci.border = THIN

        # Auto-height hint (approximate: 1 doc = 18pt)
        n_lines = max(len(req.required_docs), len(tx_refs), 1)
        ws.row_dimensions[row].height = max(30, n_lines * 18)

        row += 1

    last_row = row - 1

    # ── Conditional formatting by Statut ───────────────────────────────
    data_range = f"A{DATA_START}:I{last_row}"

    # "Reçue" → green
    ws.conditional_formatting.add(
        data_range,
        FormulaRule(
            formula=[f'$F{DATA_START}="Reçue"'],
            fill=GREEN_FILL,
        ),
    )
    # "Envoyée" → light blue
    ws.conditional_formatting.add(
        data_range,
        FormulaRule(
            formula=[f'$F{DATA_START}="Envoyée"'],
            fill=BLUE_FILL,
        ),
    )
    # "Non requis" → grey
    ws.conditional_formatting.add(
        data_range,
        FormulaRule(
            formula=[f'$F{DATA_START}="Non requis"'],
            fill=ALT_FILL,
        ),
    )
    # "À envoyer" → light orange
    ws.conditional_formatting.add(
        data_range,
        FormulaRule(
            formula=[f'$F{DATA_START}="À envoyer"'],
            fill=YELLOW_FILL,
        ),
    )

    # Set print area for easy printing/emailing
    ws.print_area = f"A1:I{last_row}"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1

    _set_widths(ws, {
        "A": 5, "B": 30, "C": 36, "D": 36,
        "E": 36, "F": 14, "G": 14, "H": 14, "I": 28,
    })


# ── Backward-compatible aliases (imported by tests & other modules) ────────

FLAG_TYPE_LABELS = FLAG_LABELS          # old name → new name
_format_date_short = _fmt_date          # old helper name


# ── Public API ────────────────────────────────────────────────────────────


def generate_dp_excel(result: DPAuditResult) -> bytes:
    """Generate the downpayment audit Excel workbook and return bytes."""
    wb = Workbook()

    ws_dashboard = wb.active
    ws_dashboard.title = "Tableau de bord"
    ws_analyse   = wb.create_sheet("Analyse")
    ws_demandes  = wb.create_sheet("Demandes client")

    # Analyse built first — returns row numbers needed by Dashboard formulas
    summary_rows = _fill_analyse(ws_analyse, result)
    _fill_dashboard(ws_dashboard, result, summary_rows)
    _fill_demandes(ws_demandes, result)

    # Open on Dashboard tab
    wb.active = ws_dashboard

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_dp_excel_base64(result: DPAuditResult) -> str:
    """Generate Excel and return as base64 string."""
    return base64.b64encode(generate_dp_excel(result)).decode()
