"""Generate the Excel income analysis report from extraction data."""

import base64
import io
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

from mortgage_mcp.models.bank_statement import BankStatementExtraction, DepositCategory

TEMPLATE_PATH = Path(__file__).parent.parent / "templates" / "grille_revenu_autonome.xlsx"

HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")

SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SECTION_FONT = Font(bold=True, size=11, color="1F4E79")
INPUT_FILL = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")
INPUT_FONT = Font(color="9E9E9E", italic=True)
ALT_ROW_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

CURRENCY_FORMAT = '#,##0.00 $'
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Category row fill colors for Dépôts sheet
_CAT_FILLS = {
    DepositCategory.BUSINESS_INCOME: PatternFill(
        start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"
    ),
    DepositCategory.PERSONAL_TRANSFER: PatternFill(
        start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"
    ),
    DepositCategory.GOVERNMENT: PatternFill(
        start_color="F3E5F5", end_color="F3E5F5", fill_type="solid"
    ),
    DepositCategory.REFUND: PatternFill(
        start_color="FFF8E1", end_color="FFF8E1", fill_type="solid"
    ),
}


def _apply_header_style(ws, row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def _section_title(ws, row: int, text: str, max_col: int = 3) -> None:
    """Write a styled section header row (merges cols 1..max_col)."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL


def generate_excel(extraction: BankStatementExtraction) -> bytes:
    """Populate the Excel template with extraction data and return bytes.

    If the template file exists, it is used as-is. Otherwise, sheets are
    created from scratch with proper formatting.
    """
    if TEMPLATE_PATH.exists():
        wb = load_workbook(TEMPLATE_PATH)
    else:
        wb = _create_workbook_from_scratch()

    _fill_resume(wb, extraction)
    _fill_monthly(wb, extraction)
    _fill_deposits(wb, extraction)
    _fill_withdrawals(wb, extraction)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_excel_base64(extraction: BankStatementExtraction) -> str:
    """Generate Excel and return as base64 string."""
    return base64.b64encode(generate_excel(extraction)).decode()


def _create_workbook_from_scratch():
    """Create a new workbook with the three required sheets."""
    from openpyxl import Workbook

    wb = Workbook()
    ws_resume = wb.active
    ws_resume.title = "Resume"
    wb.create_sheet("Detail mensuel")
    wb.create_sheet("Depots")
    wb.create_sheet("Retraits")
    return wb


def _fill_resume(wb, extraction: BankStatementExtraction) -> None:
    """Fill the Resume (summary) sheet."""
    ws = wb["Resume"]

    # Remove any existing merged cells, then clear values
    for merge_range in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merge_range))
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.value = None

    info = extraction.account_info
    n = len(extraction.monthly_breakdown)
    # Row positions in Detail mensuel for formula references
    dm_total_row = n + 2   # TOTAL (SUM) row
    dm_avg_row   = n + 3   # MOYENNE (AVERAGE) row

    current_row = 1

    # ── Title ─────────────────────────────────────────────────────────────────
    ws.cell(
        row=current_row, column=1,
        value="Grille d'analyse — Revenu de travailleur autonome",
    ).font = Font(bold=True, size=14)
    current_row += 2  # blank → 3

    # ── Borrower info (rows 3–7) ───────────────────────────────────────────
    borrower_data = [
        ("Titulaire du compte:", info.account_holder),
        ("Institution financière:", info.institution),
        ("Numéro de compte:", f"***{info.account_number_last4}" if info.account_number_last4 else "N/D"),
        ("Période analysée:", f"{info.statement_period_start} au {info.statement_period_end}"),
        ("Nombre de mois:", extraction.months_covered),
    ]
    for label, value in borrower_data:
        ws.cell(row=current_row, column=1, value=label).font = HEADER_FONT
        ws.cell(row=current_row, column=2, value=value)
        current_row += 1
    # current_row = 8 after 5 items (3+5)
    current_row += 1  # blank → 9

    # ── Sommaire financier (rows 9–14) ────────────────────────────────────
    ws.cell(row=current_row, column=1, value="Sommaire financier").font = Font(bold=True, size=12)
    current_row += 1  # → 10

    # Row 10: Dépôts totaux (static)
    ws.cell(row=current_row, column=1, value="Dépôts totaux:").font = HEADER_FONT
    ws.cell(row=current_row, column=2, value=extraction.total_deposits).number_format = CURRENCY_FORMAT
    current_row += 1  # → 11

    # Row 11: Revenu d'affaires total — formula → Detail mensuel TOTAL row col C
    ws.cell(row=current_row, column=1, value="Revenu d'affaires total:").font = HEADER_FONT
    cell = ws.cell(row=current_row, column=2, value=f"='Detail mensuel'!C{dm_total_row}")
    cell.number_format = CURRENCY_FORMAT
    current_row += 1  # → 12

    # Row 12: Retraits totaux (static)
    ws.cell(row=current_row, column=1, value="Retraits totaux:").font = HEADER_FONT
    ws.cell(row=current_row, column=2, value=extraction.total_withdrawals).number_format = CURRENCY_FORMAT
    current_row += 1  # → 13

    # Row 13: Revenu mensuel moyen — formula → Detail mensuel MOYENNE row col C
    ws.cell(row=current_row, column=1, value="Revenu mensuel moyen (affaires):").font = HEADER_FONT
    cell = ws.cell(row=current_row, column=2, value=f"='Detail mensuel'!C{dm_avg_row}")
    cell.number_format = CURRENCY_FORMAT
    current_row += 1  # → 14

    # Row 14: Revenu annualisé — formula = B13 * 12
    ws.cell(row=current_row, column=1, value="Revenu annualisé (affaires):").font = HEADER_FONT
    cell = ws.cell(row=current_row, column=2, value=f"=B{current_row - 1}*12")
    cell.number_format = CURRENCY_FORMAT
    current_row += 2  # blank → 16

    # ── Risk indicators (conditional) ─────────────────────────────────────
    if extraction.nsf_events:
        ws.cell(row=current_row, column=1, value="Indicateurs de risque").font = Font(bold=True, size=12)
        current_row += 1
        ws.cell(row=current_row, column=1, value="Événements NSF/découverts:").font = HEADER_FONT
        ws.cell(row=current_row, column=2, value=len(extraction.nsf_events))
        current_row += 1
        ws.cell(row=current_row, column=1, value="Frais NSF totaux:").font = HEADER_FONT
        ws.cell(row=current_row, column=2, value=extraction.nsf_total_fees).number_format = CURRENCY_FORMAT
        current_row += 2

    # ── Recurring obligations (conditional) ───────────────────────────────
    if extraction.recurring_obligations:
        ws.cell(row=current_row, column=1, value="Obligations récurrentes détectées").font = Font(bold=True, size=12)
        current_row += 1
        for col, header in enumerate(["Bénéficiaire", "Montant mensuel", "Type"], 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = HEADER_FONT
            cell.border = THIN_BORDER
        current_row += 1
        for obligation in extraction.recurring_obligations:
            ws.cell(row=current_row, column=1, value=obligation.payee).border = THIN_BORDER
            cell_amt = ws.cell(row=current_row, column=2, value=obligation.monthly_amount)
            cell_amt.number_format = CURRENCY_FORMAT
            cell_amt.border = THIN_BORDER
            ws.cell(row=current_row, column=3, value=obligation.category).border = THIN_BORDER
            current_row += 1
        ws.cell(row=current_row, column=1, value="Total obligations mensuelles:").font = HEADER_FONT
        ws.cell(row=current_row, column=2, value=extraction.total_monthly_obligations).number_format = CURRENCY_FORMAT
        current_row += 2

    # ── Confidence notes (conditional) ────────────────────────────────────
    if extraction.confidence_notes:
        ws.cell(row=current_row, column=1, value="Notes et observations").font = Font(bold=True, size=12)
        current_row += 1
        for note in extraction.confidence_notes:
            ws.cell(row=current_row, column=1, value=f"• {note}")
            current_row += 1
        current_row += 1

    # ── Broker sections ───────────────────────────────────────────────────

    # Dossier info
    _section_title(ws, current_row, "Informations du dossier — À compléter par le courtier", max_col=3)
    current_row += 1
    for label in ["Prêteur / programme:", "Co-emprunteur:", "Date de clôture:", "Numéro de dossier:"]:
        ws.cell(row=current_row, column=1, value=label).font = HEADER_FONT
        cell = ws.cell(row=current_row, column=2)
        cell.fill = INPUT_FILL
        cell.border = THIN_BORDER
        current_row += 1
    current_row += 1  # blank

    # Revenu qualifiable
    _section_title(ws, current_row, "Revenu qualifiable", max_col=3)
    current_row += 1
    rev_annuel_row = current_row
    ws.cell(row=current_row, column=1, value="Revenu brut annualisé (IA):").font = HEADER_FONT
    cell = ws.cell(row=current_row, column=2, value="=B14")
    cell.number_format = CURRENCY_FORMAT
    current_row += 1
    ajust_row = current_row
    ws.cell(row=current_row, column=1, value="Ajustement courtier (%):").font = HEADER_FONT
    cell = ws.cell(row=current_row, column=2, value=1.0)
    cell.number_format = "0%"
    cell.fill = INPUT_FILL
    cell.border = THIN_BORDER
    current_row += 1
    ws.cell(row=current_row, column=1, value="Revenu qualifiable retenu:").font = HEADER_FONT
    cell = ws.cell(row=current_row, column=2, value=f"=B{rev_annuel_row}*B{ajust_row}")
    cell.number_format = CURRENCY_FORMAT
    current_row += 2  # blank

    # Signature block
    _section_title(ws, current_row, "Attestation du courtier", max_col=3)
    current_row += 1
    for label in ["Courtier:", "Date:", "Signature:"]:
        ws.cell(row=current_row, column=1, value=label).font = HEADER_FONT
        cell = ws.cell(row=current_row, column=2)
        cell.fill = INPUT_FILL
        cell.border = THIN_BORDER
        current_row += 1

    # Frozen pane + column widths
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 20


def _fill_monthly(wb, extraction: BankStatementExtraction) -> None:
    """Fill the Detail mensuel (monthly breakdown) sheet."""
    ws = wb["Detail mensuel"]

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.value = None

    headers = [
        "Mois", "Dépôts bruts", "Dépôts affaires", "Transferts personnels",
        "Gouvernement", "Remboursements", "Prêts/crédit", "Autres",
        "Retraits", "Revenu net", "Nb dépôts",
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    _apply_header_style(ws, 1, len(headers))

    currency_cols = len(headers) - 1  # all except last (Nb dépôts)

    for i, month in enumerate(extraction.monthly_breakdown, start=2):
        dr = i  # current data row
        # Col C: SUMPRODUCT formula — sums deposits where Inclure="O" for this month
        sumproduct_c = (
            f"=SUMPRODUCT("
            f"(LEFT(Depots!$A$2:$A$2000,7)=A{dr})"
            f"*(Depots!$F$2:$F$2000=\"O\")"
            f"*Depots!$D$2:$D$2000)"
        )
        row_data = [
            month.month,               # A: Mois
            month.total_deposits,      # B: Dépôts bruts (static)
            sumproduct_c,              # C: Dépôts affaires (formula)
            month.personal_transfers,  # D: static
            month.government_deposits, # E: static
            month.refund_deposits,     # F: static
            month.loan_credit_deposits,# G: static
            month.other_deposits,      # H: static
            month.total_withdrawals,   # I: static
            f"=C{dr}-I{dr}",           # J: Revenu net (formula)
            month.deposit_count,       # K: static
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=dr, column=col, value=val)
            cell.border = THIN_BORDER
            if 2 <= col <= currency_cols:
                cell.number_format = CURRENCY_FORMAT
            # Alternating row color (rows 3, 5, 7…)
            if dr % 2 == 1:
                cell.fill = ALT_ROW_FILL

    # Summary rows
    n = len(extraction.monthly_breakdown)
    if n > 0:
        summary_row = n + 2
        last_data_row = n + 1

        ws.cell(row=summary_row, column=1, value="TOTAL").font = HEADER_FONT
        for col in range(2, currency_cols + 1):
            col_letter = chr(64 + col)
            cell = ws.cell(
                row=summary_row, column=col,
                value=f"=SUM({col_letter}2:{col_letter}{last_data_row})"
            )
            cell.number_format = CURRENCY_FORMAT
            cell.font = HEADER_FONT
            cell.border = THIN_BORDER

        avg_row = summary_row + 1
        ws.cell(row=avg_row, column=1, value="MOYENNE").font = HEADER_FONT
        for col in range(2, currency_cols + 1):
            col_letter = chr(64 + col)
            cell = ws.cell(
                row=avg_row, column=col,
                value=f"=AVERAGE({col_letter}2:{col_letter}{last_data_row})"
            )
            cell.number_format = CURRENCY_FORMAT
            cell.font = HEADER_FONT
            cell.border = THIN_BORDER

    # Frozen pane + column widths
    ws.freeze_panes = "A2"
    widths = [12, 16, 16, 20, 16, 18, 16, 14, 16, 16, 12]
    for i, w in enumerate(widths):
        ws.column_dimensions[chr(65 + i)].width = w


def _fill_deposits(wb, extraction: BankStatementExtraction) -> None:
    """Fill the Depots (all deposits) sheet."""
    ws = wb["Depots"]

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.value = None

    headers = [
        "Date", "Compte", "Description", "Montant", "Catégorie",
        "Inclure (O/N)", "Explication courtier",
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    _apply_header_style(ws, 1, len(headers))

    # Dropdown data validation for "Inclure" column F
    dv = DataValidation(type="list", formula1='"O,N"', allow_blank=False)
    dv.add("F2:F2000")
    ws.add_data_validation(dv)

    row_num = 2
    for month in extraction.monthly_breakdown:
        for dep in month.deposits:
            cat = dep.category
            inclure = "O" if cat == DepositCategory.BUSINESS_INCOME else "N"
            cat_fill = _CAT_FILLS.get(cat)

            data = [
                dep.date, dep.account, dep.description,
                dep.amount, dep.category.value, inclure,
            ]
            for col, val in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col, value=val)
                cell.border = THIN_BORDER
                if col == 4:
                    cell.number_format = CURRENCY_FORMAT
                if col == 6:
                    cell.alignment = Alignment(horizontal="center")
                # Category-based fill on cols A–E
                if cat_fill and col <= 5:
                    cell.fill = cat_fill

            # Col G: Explication courtier (broker input, pale yellow)
            cell_g = ws.cell(row=row_num, column=7)
            cell_g.fill = INPUT_FILL
            cell_g.border = THIN_BORDER

            row_num += 1

    # AutoFilter + frozen pane
    ws.auto_filter.ref = f"A1:G{max(row_num - 1, 1)}"
    ws.freeze_panes = "A2"

    # Column widths
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 35


def _fill_withdrawals(wb, extraction: BankStatementExtraction) -> None:
    """Fill the Retraits (all withdrawals) sheet."""
    if "Retraits" not in wb.sheetnames:
        wb.create_sheet("Retraits")
    ws = wb["Retraits"]

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.value = None

    headers = ["Date", "Compte", "Description", "Montant", "Catégorie", "Commentaire courtier"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    _apply_header_style(ws, 1, len(headers))

    row_num = 2
    for month in extraction.monthly_breakdown:
        for wd in month.withdrawals:
            ws.cell(row=row_num, column=1, value=wd.date).border = THIN_BORDER
            ws.cell(row=row_num, column=2, value=wd.account).border = THIN_BORDER
            ws.cell(row=row_num, column=3, value=wd.description).border = THIN_BORDER
            cell_amt = ws.cell(row=row_num, column=4, value=wd.amount)
            cell_amt.number_format = CURRENCY_FORMAT
            cell_amt.border = THIN_BORDER
            ws.cell(row=row_num, column=5, value=wd.category).border = THIN_BORDER
            # Col F: Commentaire courtier (broker input)
            cell_f = ws.cell(row=row_num, column=6)
            cell_f.fill = INPUT_FILL
            cell_f.border = THIN_BORDER
            row_num += 1

    # AutoFilter + frozen pane
    ws.auto_filter.ref = f"A1:F{max(row_num - 1, 1)}"
    ws.freeze_panes = "A2"

    # Column widths
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 35
